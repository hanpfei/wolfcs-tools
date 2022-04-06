#!/use/bin/env python

import openpyxl


def main():
    file_path = "/home/hanpfei/Week7完整成绩单.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sh = wb['Sheet1']

    sh.cell(row=1, column=1)

    rows_data = list(sh.rows)

    titles = [title_ce.value for title_ce in rows_data[0]]
    print(titles)

    correct_answer_line = rows_data[1]
    correct_answers = [question_answer.value for question_answer in correct_answer_line]
    print(correct_answers)

    all_students_answer = []
    students_answer_lines = rows_data[2:]
    for students_answer_line in students_answer_lines:
        cur_student_line = [question_answer.value for question_answer in students_answer_line]
        all_students_answer.append(cur_student_line)

    # print(all_students_answer)

    analyze_result = []
    answer_start_column = 4
    for column in range(len(correct_answers)):
        analyze_result.append(set({}))

        if column < answer_start_column:
            continue

        for cur_student_answers in all_students_answer:
            if cur_student_answers[column] != correct_answers[column]:
                analyze_result[column].add(cur_student_answers[column])

    print(analyze_result)

    output_wb = openpyxl.Workbook()
    o_sh = output_wb.create_sheet('Sheet')

    for column in range(len(titles)):
        cur_column = column + 1
        o_sh.cell(row=1, column=cur_column, value=titles[column])
        o_sh.cell(row=2, column=cur_column, value=correct_answers[column])

    for column in range(answer_start_column, len(titles)):
        cur_row = 3
        cur_column = column + 1
        for wrong_answer in analyze_result[column]:
            o_sh.cell(row=cur_row, column=cur_column, value=wrong_answer)
            cur_row = cur_row + 1

    output_wb.save('/home/hanpfei/output.xlsx')


if __name__ == "__main__":
    main()